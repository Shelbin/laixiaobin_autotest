# -*- coding: utf-8 -*-
# @Author: nickel
# @File:create_testcase.py
# @Time: 2021/11/6 18:44

from openpyxl import Workbook

def create_testcase():
    '''
    编写接口测试用例
    :return:
    '''
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.title = "testcase"
    # 插入测试用例的表头
    case_title = ["用例编号", "请求方式", "URL", "预期结果", "实际结果"]
    sheet.append(case_title)

    # 用单元格插入数据的方式，插入一条测试用例
    sheet["A2"] = 1
    sheet["B2"] = "GET"
    sheet["C2"] = "https://www.baidu.com"
    sheet["D2"] = "baidu"

    # 使用循环的方式插入5条测试用例
    for row in range(3,8):
        for col in range(1, 5):
            if col == 1:
                sheet.cell(row=row, column=col, value=row-1)
            elif col == 2:
                sheet.cell(row=row, column=col, value="GET")
            elif col == 3:
                sheet.cell(row=row, column=col, value="https://www.baidu.com")
            else:
                sheet.cell(row=row, column=col, value="baidu")
                if row == 4:
                    sheet.cell(row=row, column=col, value="xxx")

    wb.save(filename="cases.xlsx")
    wb.close()

if __name__ == "__main__":
    create_testcase()
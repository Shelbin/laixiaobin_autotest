# -*- coding: utf-8 -*-
# @Author: nickel
# @File:read_testcase.py
# @Time: 2021/11/7 10:19
from openpyxl import load_workbook

def read_testcase():
    '''
    读取测试用例
    :return:
    '''
    wb = load_workbook('cases.xlsx')
    sheet = wb.worksheets[0]
    # 获取测试用例的行数
    case_num = sheet.max_row

    # 接口数据列表
    interface_list = []
    # 循环获取表格中的数据
    for row in range(2, case_num+1):
        r1 = []
        for col in range(2, 5):
            if col == 2:
                method = sheet.cell(row,col).value
                r1.append(method)
            elif col == 3:
                url = sheet.cell(row,col).value
                r1.append(url)
            else:
                expected = sheet.cell(row,col).value
                r1.append(expected)
        interface_list.append(r1)
    return interface_list

if __name__ == '__main__':
    print(read_testcase())

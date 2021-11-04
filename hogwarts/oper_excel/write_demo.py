from openpyxl import Workbook

wb = Workbook()

ws = wb.active

# 创建新的sheet，直接在末尾添加
ws1 = wb.create_sheet(title="Mysheet")

# 在指定index处添加
ws2 = wb.create_sheet(title="Mysheet2", index=0)

# sheet命名
ws.title = "New title"

# 调整背景色
ws.sheet_properties.tabColor = "1072BA"

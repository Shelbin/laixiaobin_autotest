from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 实例化一个工作簿对象
wb = Workbook()

# 定义 excel 文件名
dest_filename = 'empty_book.xlsx'

# 选取了当前工作的sheet对象
ws1 = wb.active

# 为 sheet 命名
ws1.title = "range names"
for row in range(1,40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title='Pi')
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10,20):
    for col in range(27,54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col))) # 得到表格列的字母编号

print(ws3['AA10'].value)

wb.save(filename=dest_filename)



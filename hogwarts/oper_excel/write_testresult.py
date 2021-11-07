# -*- coding: utf-8 -*-
# @Author: nickel
# @File:write_testresult.py
# @Time: 2021/11/7 10:44

from openpyxl import load_workbook

def write_result(test_result):
    wb = load_workbook('cases.xlsx')
    sheet = wb.worksheets[0]

    for row in range(2, 8):
        sheet.cell(row=row, column=5, value=test_result[row-2])

    wb.save("cases.xlsx")
    wb.close()